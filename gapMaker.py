#! python3
# gapMakerExcel.py

import os, openpyxl, logging, sys

from openpyxl.cell import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Style

logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')
# logging.disable(logging.CRITICAL)

path = '.\\excel\\'
if not os.path.exists(path):
    os.makedirs(path)
os.chdir(path)

wb = openpyxl.load_workbook('table5x5.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

wbNew = openpyxl.Workbook()
sheetNew = wbNew.get_sheet_by_name('Sheet')

startLineNumber = input('start line: ')
blankLinesNumber = input('num. of blank lines: ')

for row in range(sheet.max_row):
    col = row + 1
    copyValue = get_column_letter(col) + str((row + 1))
    logging.debug(copyValue)
    sheet.cell(row=2, column=5)

wbNew.save('newTable.xlsx')