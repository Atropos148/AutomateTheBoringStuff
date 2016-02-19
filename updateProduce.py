#! python3
# updateProduce.py

import os, openpyxl, logging

from openpyxl.cell import get_column_letter, column_index_from_string

logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')
# logging.disable(logging.CRITICAL)

path = '.\\excel\\'
if not os.path.exists(path):
    os.makedirs(path)
os.chdir(path)

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')
# produce and updated price
PRICE_UPDATE = {'Garlic': 3.07,
                'Celery': 1.19,
                'Lemon': 1.27}
incomeTotal = 0

# loop through rows and update price

for rowNum in range(2, sheet.max_row):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATE:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATE[produceName]

wb.save('updatedProduceSales.xlsx')
