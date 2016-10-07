#! python3
# timeSheet.py - Helps keep track of check-in and check-out of guests


import os, time, openpyxl, datetime

from openpyxl import Workbook, load_workbook
from datetime import datetime 

print('Welcome...')

path = '.\sheets'
if not os.path.exists(path):
    os.makedirs(path)
os.chdir(path)

sheetPath = 'timeSheet.xlsx'
if os.path.isfile(sheetPath) == True:
    print('Workbook found...')
else:
    print('Creating workbook...')
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'timeSheet'
    wb.save(filename = sheetPath)

print('Opening workbook...')
wb = load_workbook(sheetPath)
print('Workbook opened...')
sheet = wb.get_sheet_by_name('timeSheet')
print('Sheet found...')

timestamp = datetime.now()
print('Guest name:')
guestName = input('')

nextRow = sheet.max_row + 1

sheet = wb.active

sheet['A' + str(nextRow)] = guestName
sheet['B' + str(nextRow)] = timestamp

print('Saving Data...')
wb.save('timeSheet.xlsx')
print('Done...')
