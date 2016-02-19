#! python3
# updateProduce.py

import os, openpyxl, logging, sys

from openpyxl.cell import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Style

logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')
# logging.disable(logging.CRITICAL)

path = '.\\excel\\'
if not os.path.exists(path):
    os.makedirs(path)
os.chdir(path)

wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')

iNum = input()

iCol = 1
iRow = 1
rowNum = 2

boldFont = Font(bold=True)

for col in range(2, int(iNum)+2):
    sheet.cell(row=1, column=col).font = boldFont
    sheet.cell(row=1, column=col).value = iCol
    iCol += 1

    for rowNum in range(2, int(iNum)+2):
        topVal = get_column_letter(col) + '1'
        sideVal = 'A' + str(rowNum)
        multiplyText = '=' + topVal + '*' + sideVal
        sheet.cell(row=rowNum, column=col).value = multiplyText
        rowNum += 1

for rowNum in range(2, int(iNum)+2):
    sheet.cell(row=rowNum, column=1).font = boldFont
    sheet.cell(row=rowNum, column=1).value = iRow
    iRow += 1

tableName = 'table' + iNum + 'x' + iNum + '.xlsx'
wb.save(tableName)
print('Done')
